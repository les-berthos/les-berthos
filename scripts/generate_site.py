#!/usr/bin/env python3
"""
Génère le site statique (dossier /docs) à partir de data/Genealogie_MEME.xlsx.
Ne modifie jamais le xlsx. À relancer à chaque mise à jour du tableau
(en pratique : automatique via .github/workflows/build-site.yml).
"""
import re
from pathlib import Path
import pandas as pd
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "data" / "Genealogie_MEME.xlsx"
DOCS = ROOT / "docs"
PERSONNES_DIR = DOCS / "personnes"

CONTACT_EMAIL = "FamBertho@proton.me"
SITE_TITLE = "Les Berthos"
SITE_TAGLINE = "Une enquête de famille, page après page — mise à jour au fil des découvertes."


def slug(text: str) -> str:
    text = str(text).strip()
    text = re.sub(r"[^\w\-]+", "-", text, flags=re.UNICODE)
    return re.sub(r"-+", "-", text).strip("-").upper()


def is_blank(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    return s == "" or s.lower() == "nan" or s == "?"


def split_list(v):
    if is_blank(v):
        return []
    return [x.strip() for x in re.split(r"[;,:\n]", str(v)) if x.strip()]


def split_urls(v):
    """Comme split_list mais sans le ':' comme séparateur, pour ne pas couper les liens http(s)://"""
    if is_blank(v):
        return []
    return [x.strip() for x in re.split(r"[;,\n]", str(v)) if x.strip()]


MOIS_FR = ["janvier", "février", "mars", "avril", "mai", "juin", "juillet",
           "août", "septembre", "octobre", "novembre", "décembre"]


def format_date_fr(raw):
    """Reformate une date en 'DD mois AAAA'. Si non reconnue, renvoie le texte tel quel."""
    if is_blank(raw):
        return raw
    s = str(raw).strip()
    # Cas Excel qui a stocké une vraie date (ex. '1913-06-16 00:00:00' ou '1913-06-16')
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12:
            return f"{d} {MOIS_FR[mo-1]} {y}"
    # Cas texte 'JJ/MM/AAAA'
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12:
            return f"{d} {MOIS_FR[mo-1]} {y}"
    return s  # texte libre ('vers 1895', '1901', etc.) affiché tel quel


def field_status(raw):
    """Renvoie (texte_affiché, statut) où statut est 'confirme' / 'verifier' / 'manque'.
    Un '?' en fin de case = à vérifier. Case vide = à enquêter."""
    if is_blank(raw):
        return "", "manque"
    s = str(raw).strip()
    if s.endswith("?"):
        return format_date_fr(s[:-1].strip()), "verifier"
    return format_date_fr(s), "confirme"


def mini_stamp(status):
    label = {"confirme": "confirmé", "verifier": "à confirmer", "manque": "à enquêter"}[status]
    return f'<span class="field-stamp {status}">{label}</span>'


def compute_branches(personnes: pd.DataFrame):
    """Calcule automatiquement, pour chaque personne, si elle est du côté MEME ou du côté
    PEPE, en remontant/descendant les liens Père/Mère/Fratrie depuis MEME et son conjoint.
    Renvoie (branch, meme_id, pepe_id)."""
    idx = by_id_index(personnes)
    branch = {}
    meme_id = None
    for pid, row in idx.items():
        if str(get_lien(row)).strip().lower() == "meme":
            meme_id = pid
            break
    if not meme_id:
        return branch, None, None
    branch[meme_id] = "MEME"
    conjoint_col = find_col(idx[meme_id], CONJOINT_COLONNES)
    pepe_id = str(idx[meme_id].get(conjoint_col, "")).strip()
    pepe_id = split_list(pepe_id)[0] if split_list(pepe_id) else ""
    if pepe_id and pepe_id in idx:
        branch[pepe_id] = "PEPE"

    changed = True
    passes = 0
    while changed and passes < len(personnes) + 2:
        changed = False
        passes += 1
        for pid, row in idx.items():
            if pid in branch:
                continue
            for kid in find_children(pid, personnes):
                kid_id = str(kid.get("ID Personne", "")).strip()
                if kid_id in branch:
                    branch[pid] = branch[kid_id]
                    changed = True
                    break
            if pid in branch:
                continue
            for sib_id in split_list(row.get("Fratrie (ID Personne)", "")):
                if sib_id in branch:
                    branch[pid] = branch[sib_id]
                    changed = True
                    break
            if pid in branch:
                continue
            # hérite de la branche du père ou de la mère, sauf si le parent est
            # MEME ou PEPE lui-même (leurs enfants communs ne sont ni l'un ni l'autre)
            pere_col = find_col(row, PERE_COLONNES)
            mere_col = find_col(row, MERE_COLONNES)
            for parent_id in (str(row.get(pere_col, "")).strip(), str(row.get(mere_col, "")).strip()):
                if parent_id and parent_id not in (meme_id, pepe_id) and parent_id in branch:
                    branch[pid] = branch[parent_id]
                    changed = True
                    break
    return branch, meme_id, pepe_id


def load_data():
    personnes = pd.read_excel(XLSX, sheet_name="Registre des personnes", dtype=str)
    personnes = personnes.dropna(how="all")
    if "ID Personne" in personnes.columns:
        personnes = personnes[~personnes["ID Personne"].astype(str).str.contains("EXEMPLE", na=False)]
    personnes = personnes.fillna("")
    return personnes


def load_documents():
    try:
        wb = openpyxl.load_workbook(XLSX, data_only=True)
    except Exception:
        return pd.DataFrame()
    sheet_name = None
    for candidate in ("Registre des documents", "Feuille1"):
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if not sheet_name:
        return pd.DataFrame()
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    if "ID Document" not in headers:
        return pd.DataFrame()
    id_col = headers.index("ID Document") + 1

    rows = []
    for r in range(2, ws.max_row + 1):
        rowvals = {}
        has_content = False
        for ci, h in enumerate(headers, start=1):
            if not h:
                continue
            v = ws.cell(row=r, column=ci).value
            if v not in (None, ""):
                has_content = True
            rowvals[str(h)] = "" if v is None else str(v)
        if not has_content:
            continue
        link_cell = ws.cell(row=r, column=id_col)
        rowvals["_lien_hyperlien"] = link_cell.hyperlink.target if link_cell.hyperlink else ""
        rows.append(rowvals)
    return pd.DataFrame(rows)


def get_doc_link(row):
    """Cherche le lien du document, quel que soit le nom de colonne utilisé, avec
    en dernier recours l'hyperlien posé directement sur la cellule ID Document."""
    for key in ("Photo du document", "Lien", "Lien du document", "Lien Proton", "Scan"):
        v = row.get(key, "")
        if not is_blank(v):
            return v
    return row.get("_lien_hyperlien", "")


PERSONNES_MENTIONNEES_COLONNES = ["Personnes mentionnées (ID)", "Personnes mentionnées", "Autres personnes mentionnées"]


def get_personnes_mentionnees(doc_row):
    ids = []
    for col in PERSONNES_MENTIONNEES_COLONNES:
        ids += split_list(doc_row.get(col, ""))
    return ids


def find_documents_for_person(pid, documents: pd.DataFrame):
    if documents.empty:
        return []
    pid = str(pid).strip()
    matches = []
    for _, row in documents.iterrows():
        if pid in get_personnes_mentionnees(row):
            matches.append(row)
    return matches


def render_documents_block(pid, documents, idx, title="Sources", compact=False):
    docs = find_documents_for_person(pid, documents)
    if not docs:
        return ""

    def person_label(other_pid):
        other_pid = other_pid.strip()
        if other_pid == str(pid).strip():
            return None
        row = idx.get(other_pid)
        label = person_display_name(row) if row is not None else other_pid
        return f'<a href="#{slug(other_pid)}">{label}</a>' if row is not None else label

    chips = []
    for d in docs:
        nom_doc = d.get("Nom du doc") or d.get("Type de document") or d.get("ID Document") or "Document"
        lien = get_doc_link(d)
        lien_html = f'<a href="{lien}" target="_blank" rel="noopener">voir</a>' if not is_blank(lien) else "<span class='empty'>pas encore de lien</span>"
        if compact:
            chips.append(f"<div class='doc-chip'><strong>{nom_doc}</strong> — {lien_html}</div>")
            continue
        lieu = d.get("Rangement physique", "")
        autres_ids = set(get_personnes_mentionnees(d))
        autres_ids.discard(str(pid).strip())
        autres = [person_label(a) for a in autres_ids]
        autres = [a for a in autres if a]
        chips.append(f"""
        <div class="doc-chip">
          <strong>{nom_doc}</strong> — {lien_html}
          {f"<span class='meta'>Chez : {lieu}</span>" if not is_blank(lieu) else ""}
          {f"<span class='meta'>Avec : {', '.join(autres)}</span>" if autres else ""}
        </div>""")
    return f"<h3>{title}</h3><div class='doc-list'>{''.join(chips)}</div>"


def person_display_name(row) -> str:
    prenom = row.get("Prénom(s)", "")
    nom = row.get("Nom", "")
    if is_blank(prenom) and is_blank(nom):
        return row.get("ID Personne", "Personne inconnue")
    return f"{'' if is_blank(prenom) else prenom} {'' if is_blank(nom) else nom}".strip()


LIEN_COLONNES = ["Lien avec la MEME", "Lien avec PEPE ou MEME", "Lien avec la Meme"]


def get_lien(row):
    for col in LIEN_COLONNES:
        v = row.get(col, "")
        if not is_blank(v):
            return v
    return ""


def by_id_index(personnes: pd.DataFrame):
    idx = {}
    for _, row in personnes.iterrows():
        pid = str(row.get("ID Personne", "")).strip()
        if pid:
            idx[pid] = row
    return idx


def find_children(pid, personnes: pd.DataFrame):
    kids = []
    for _, row in personnes.iterrows():
        pere = str(row.get("Père (ID Personne)", "")).strip()
        mere = str(row.get("Mère (ID Personne)", "")).strip()
        if pere == pid or mere == pid:
            kids.append(row)
    return kids


def compute_leads(personnes: pd.DataFrame):
    leads = []
    key_fields = ["Date de naissance", "Lieu de naissance", "Date de décès", "Lieu de décès"]
    for _, row in personnes.iterrows():
        pid = row.get("ID Personne", "")
        name = person_display_name(row)
        piste = row.get("Piste de recherche", "")
        if not is_blank(piste):
            leads.append({"id": pid, "name": name, "text": piste})
        else:
            manquants = [f for f in key_fields if is_blank(row.get(f, ""))]
            if manquants and row.get("Statut (confirmé / hypothèse)", "") != "confirmé":
                leads.append({
                    "id": pid, "name": name,
                    "text": f"Champs encore inconnus : {', '.join(manquants)}."
                })
    return leads


CSS = """
:root{
  --paper:#efe7d8; --paper-dark:#e4d9c3; --ink:#3a2e22; --ink-soft:#6b5c48;
  --wax:#8c2f39; --brass:#b8863b; --sage:#5b7065; --line:#c9bda3; --cream:#f6efdd;
  --meme:#a8455c; --pepe:#3f6a78;
}
*{box-sizing:border-box}
body{
  margin:0; color:var(--ink);
  font-family:'Public Sans', Arial, sans-serif; font-size:18px; line-height:1.55;
  background-image:
    radial-gradient(var(--paper-dark) 0.6px, transparent 0.6px),
    linear-gradient(rgba(239,231,216,.87), rgba(239,231,216,.87)),
    url('https://i.postimg.cc/vmTKF4ZQ/arriere-plan-Bertho.png');
  background-size: 14px 14px, cover, cover;
  background-position: 0 0, center, center;
  background-repeat: repeat, no-repeat, no-repeat;
  background-attachment: scroll, fixed, fixed;
}
.hero{
  position:relative;
  background-image: linear-gradient(180deg, rgba(20,14,8,.4) 0%, rgba(20,14,8,.68) 100%);
  padding:4.5rem 1.5rem 2.2rem; text-align:center;
}
.hero h1{
  font-family:'Playfair Display', Georgia, serif; font-size:3rem; margin:0 0 .4rem;
  color:var(--cream); letter-spacing:1px; text-shadow:0 2px 10px rgba(0,0,0,.5);
}
.hero p{margin:0; color:var(--cream); opacity:.9; font-size:1.05rem; text-shadow:0 1px 6px rgba(0,0,0,.5)}
nav.site{
  display:flex; justify-content:center; gap:1.6rem; padding:.9rem 0; background:var(--paper-dark);
  border-bottom:1px solid var(--line); font-weight:600; letter-spacing:.3px; flex-wrap:wrap;
}
nav.site a{color:var(--ink); text-decoration:none; border-bottom:2px solid transparent; padding-bottom:2px}
nav.site a:hover, nav.site a:focus{border-bottom-color:var(--wax)}
main{max-width:880px; margin:0 auto; padding:2.2rem 1.5rem 4rem}
h2{font-family:'Playfair Display', Georgia, serif; color:var(--wax); font-size:1.7rem; margin-top:2.2rem}
.card{
  background:#faf6ec; border:1px solid var(--line); border-radius:3px;
  padding:1.4rem 1.6rem; margin:1.1rem 0; box-shadow:0 1px 0 var(--line);
}
.card a.name{font-family:'Playfair Display', Georgia, serif; font-size:1.3rem; color:var(--ink); text-decoration:none}
.card a.name:hover{color:var(--wax)}
.stamp{
  display:inline-block; font-family:'Special Elite', 'Courier New', monospace; font-size:.72rem;
  text-transform:uppercase; letter-spacing:1.5px; padding:.25rem .55rem; border:2px solid; border-radius:2px;
  transform:rotate(-3deg); margin-left:.6rem; vertical-align:middle;
}
.stamp.confirme{color:var(--sage); border-color:var(--sage)}
.stamp.hypothese{color:var(--brass); border-color:var(--brass)}
.stamp.inconnu{color:var(--wax); border-color:var(--wax)}
.meta{color:var(--ink-soft); font-size:.95rem; margin:.3rem 0 0}
.histoire{margin-top:.7rem}
.piste{
  background:#f3e7c9; border-left:4px solid var(--brass); padding:.7rem 1rem; margin-top:.8rem;
  font-size:.98rem;
}
.piste strong{color:var(--brass)}
.idcard{
  display:flex; gap:1.4rem; background:#faf6ec; border:2px solid var(--ink); border-radius:4px;
  padding:1.3rem; box-shadow:3px 3px 0 var(--line); flex-wrap:wrap; border-left-width:8px;
}
.idcard.branche-meme{border-left-color:var(--meme)}
.idcard.branche-pepe{border-left-color:var(--pepe)}
.legend{display:flex; gap:1.4rem; align-items:center; font-size:.9rem; color:var(--ink-soft); margin:.6rem 0 1.4rem; flex-wrap:wrap}
.legend .dot{display:inline-block; width:12px; height:12px; border-radius:50%; margin-right:.4rem; vertical-align:middle}
.legend .dot.meme{background:var(--meme)} .legend .dot.pepe{background:var(--pepe)}
.field-stamp{
  display:inline-block; font-family:'Special Elite','Courier New',monospace; font-size:.62rem;
  text-transform:uppercase; letter-spacing:1px; padding:.1rem .4rem; border-radius:2px; margin-left:.5rem;
  border:1px solid;
}
.field-stamp.confirme{color:var(--sage); border-color:var(--sage)}
.field-stamp.verifier{color:var(--brass); border-color:var(--brass)}
.field-stamp.manque{color:var(--wax); border-color:var(--wax)}
.idcard .photo{
  width:140px; height:170px; flex-shrink:0; background:var(--paper-dark); border:1px solid var(--ink-soft);
  object-fit:cover; display:flex; align-items:center; justify-content:center; color:var(--ink-soft);
  font-size:.8rem; text-align:center; padding:.5rem;
}
.idcard .infos{flex:1; min-width:220px}
.idcard .infos h2{margin:0 0 .2rem; font-size:1.5rem}
.idcard .infos .eyebrow{
  font-family:'Special Elite','Courier New',monospace; font-size:.7rem; letter-spacing:2px;
  color:var(--ink-soft); text-transform:uppercase; margin-bottom:.4rem;
}
.idcard dl{display:grid; grid-template-columns:auto 1fr; gap:.25rem .8rem; margin:.7rem 0 0; font-size:.95rem}
.idcard dt{color:var(--ink-soft)}
.idcard dd{margin:0}
.gallery{display:flex; flex-wrap:wrap; gap:.6rem; margin-top:.9rem}
.gallery a{display:block; width:90px; height:90px; overflow:hidden; border:1px solid var(--line); border-radius:3px}
.gallery img{width:100%; height:100%; object-fit:cover; display:block}
h3{font-family:'Playfair Display', Georgia, serif; color:var(--ink); font-size:1.05rem; margin:1rem 0 .4rem}
.doc-list{display:flex; flex-direction:column; gap:.5rem}
.doc-chip{
  background:var(--paper-dark); border:1px solid var(--line); border-radius:3px;
  padding:.6rem .9rem; font-size:.9rem;
}
.doc-chip .meta{display:block; margin-top:.15rem}
footer.contact{
  margin-top:3rem; padding:1.8rem; text-align:center; background:var(--paper-dark);
  border-top:1px solid var(--line);
}
footer.contact a.btn{
  display:inline-block; margin-top:.6rem; background:var(--wax); color:#faf6ec; text-decoration:none;
  padding:.7rem 1.4rem; border-radius:3px; font-weight:700;
}
footer.contact a.btn:hover{background:#732530}
.empty{color:var(--ink-soft); font-style:italic}
.back{display:inline-block; margin-bottom:1.2rem; color:var(--ink-soft); text-decoration:none}
.back:hover{color:var(--wax)}
.archive-box{background:#faf6ec; border:1px solid var(--line); border-left:4px solid var(--wax); padding:1.2rem 1.4rem; margin:1.2rem 0}
@media (max-width:600px){
  .hero h1{font-size:2.1rem} body{font-size:17px} .idcard{flex-direction:column; align-items:center; text-align:center}
  .idcard dl{grid-template-columns:1fr; text-align:center}
}
"""

FONT_LINKS = (
    '<link rel="preconnect" href="https://fonts.googleapis.com">'
    '<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@700&'
    'family=Public+Sans:wght@400;600;700&family=Special+Elite&display=swap" rel="stylesheet">'
)


def page(title, body, active="", depth=0):
    prefix = "../" * depth

    def navlink(href, label, key):
        cls = ' style="border-bottom-color:var(--wax)"' if key == active else ""
        return f'<a href="{prefix}{href}"{cls}>{label}</a>'

    nav = f"""
    <nav class="site">
      {navlink('index.html', "Accueil", 'accueil')}
      {navlink('fiches.html', "Fiches des acteurs", 'fiches')}
      {navlink('pistes.html', "Pistes de recherche", 'pistes')}
      {navlink('archives-privees.html', "Archives", 'archives')}
    </nav>"""
    if active == "accueil" and depth == 0:
        header = f"""<header class="hero"><h1>{SITE_TITLE}</h1><p>{SITE_TAGLINE}</p></header>"""
    else:
        header = f"""<header style="padding:1.6rem 1.5rem; text-align:center; border-bottom:1px solid var(--line);">
                   <a href="{prefix}index.html" style="font-family:'Playfair Display',Georgia,serif; font-size:1.5rem; color:var(--wax); text-decoration:none;">{SITE_TITLE}</a>
                 </header>"""
    return f"""<!doctype html>
<html lang="fr"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title} — {SITE_TITLE}</title>
{FONT_LINKS}
<style>{CSS}</style>
</head><body>
{header}
{nav}
<main>{body}</main>
<footer class="contact">
  <p>Une photo, une lettre, un souvenir, une correction ? Écrivez-nous, ça complète directement l'enquête.</p>
  <a class="btn" href="mailto:{CONTACT_EMAIL}?subject=Une%20piste%20pour%20l'arbre%20des%20Berthos">
    Envoyer une piste par mail</a>
</footer>
</body></html>"""


def render_idcard(row, personnes, branches, idx, link_or_text, depth, documents=None):
    pid = row.get("ID Personne", "")
    prenom = row.get("Prénom(s)", "")
    nom = row.get("Nom", "")
    nom_naissance = row.get("Nom de naissance (si différent)", "")
    sexe = str(row.get("Sexe", "")).strip().upper()
    lien_meme = get_lien(row)
    generation = row.get("Génération", "")
    branche = branches.get(str(pid).strip(), "")
    color = branch_color(branche, generation) if branche in ("MEME", "PEPE") else None
    border_style = f' style="border-left-color:{color}"' if color else ""

    # Titre : "Prénom Nom de naissance" en normal + "épouse Nom" en gras, pour les femmes mariées
    if sexe == "F" and not is_blank(nom_naissance) and nom_naissance.strip().upper() != str(nom).strip().upper():
        titre = f'<span style="font-weight:400">{prenom} {nom_naissance}</span> <span style="font-weight:800">épouse {nom}</span>'
    else:
        titre = person_display_name(row)

    photos = split_urls(row.get("Photos (liens Postimage, séparés par ;)", ""))
    portrait = photos[0] if photos else None
    gallery_photos = photos[1:] if len(photos) > 1 else []

    pere_col = find_col(row, PERE_COLONNES)
    mere_col = find_col(row, MERE_COLONNES)
    conjoint_col = find_col(row, CONJOINT_COLONNES)
    fratrie_col = find_col(row, FRATRIE_COLONNES)

    pere = link_or_text(row.get(pere_col, ""))
    pere_html = f"<dt>{label_from_header(pere_col)}</dt><dd>{pere}</dd>" if pere else ""
    mere = link_or_text(row.get(mere_col, ""))
    mere_html = f"<dt>{label_from_header(mere_col)}</dt><dd>{mere}</dd>" if mere else ""

    conjoints = [link_or_text(c) for c in split_list(row.get(conjoint_col, ""))]
    conjoints = [c for c in conjoints if c]
    conjoint_html = f"<dt>{label_from_header(conjoint_col)}</dt><dd>{', '.join(conjoints)}</dd>" if conjoints else ""

    kids = find_children(str(pid).strip(), personnes)
    kids_html = ""
    if kids:
        kid_links = [f'<a href="{"#" if depth == 0 else ""}{slug(k.get("ID Personne",""))}{".html" if depth else ""}">{person_display_name(k)}</a>' for k in kids]
        kids_html = f"<dt>Enfants</dt><dd>{', '.join(kid_links)}</dd>"

    fratrie_ids = split_list(row.get(fratrie_col, ""))
    fratrie_links = [link_or_text(f) for f in fratrie_ids]
    fratrie_links = [f for f in fratrie_links if f]
    fratrie_html = f"<dt>{label_from_header(fratrie_col)}</dt><dd>{', '.join(fratrie_links)}</dd>" if fratrie_links else ""

    nom_naissance_html = ""
    if titre == person_display_name(row) and not is_blank(nom_naissance) and nom_naissance.strip().upper() != str(nom).strip().upper():
        nom_naissance_html = f"<dt>Nom de naissance</dt><dd>{nom_naissance}</dd>"

    etat_civil_rows = [nom_naissance_html] if nom_naissance_html else []
    for label, key in [("Née/né le", "Date de naissance")]:
        txt, status = field_status(row.get(key, ""))
        etat_civil_rows.append(f"<dt>{label}</dt><dd>{txt or '—'} {mini_stamp(status)}</dd>")
    lieu_naiss = row.get("Lieu de naissance", "")
    if not is_blank(lieu_naiss):
        etat_civil_rows.append(f"<dt>Lieu de naissance</dt><dd>{lieu_naiss}</dd>")
    for label, key in [("Décédée/décédé le", "Date de décès")]:
        txt, status = field_status(row.get(key, ""))
        etat_civil_rows.append(f"<dt>{label}</dt><dd>{txt or '—'} {mini_stamp(status)}</dd>")
    lieu_deces = row.get("Lieu de décès", "")
    if not is_blank(lieu_deces):
        etat_civil_rows.append(f"<dt>Lieu de décès</dt><dd>{lieu_deces}</dd>")
    dl_html = "".join(etat_civil_rows) + conjoint_html + pere_html + mere_html + kids_html + fratrie_html

    photo_html = (
        f'<img class="photo" src="{portrait}" alt="Portrait de {person_display_name(row)}">' if portrait
        else '<div class="photo">Pas encore de photo</div>'
    )

    gallery_html = ""
    if gallery_photos:
        thumbs = "".join(
            f'<a href="{u}" target="_blank" rel="noopener"><img src="{u}" alt="Photo" loading="lazy"></a>'
            for u in gallery_photos
        )
        gallery_html = f"<div class='gallery'>{thumbs}</div>"

    docs_html = render_documents_block(pid, documents, idx, title="Sources", compact=True) if documents is not None else ""

    anchor = f'<a name="{slug(pid)}"></a>' if depth == 0 else ""
    return f"""
    {anchor}
    <div class="idcard"{border_style}>
      {photo_html}
      <div class="infos">
        <div class="eyebrow">{'' if is_blank(lien_meme) else lien_meme}</div>
        <h2>{titre} {stamp(row.get('Statut (confirmé / hypothèse)'))}</h2>
        <dl>{dl_html}</dl>
        {gallery_html}
        {docs_html}
      </div>
    </div>
    """


PERE_COLONNES = ["Père (ID Personne)", "Papa (ID Personne)"]
MERE_COLONNES = ["Mère (ID Personne)", "Maman (ID Personne)"]
CONJOINT_COLONNES = ["Conjoint(s) (ID Personne)", "Conjoint (ID Personne)"]
FRATRIE_COLONNES = ["Fratrie (ID Personne)"]


def find_col(row, candidates):
    for c in candidates:
        if c in row.index:
            return c
    return candidates[0]


def label_from_header(header):
    return re.sub(r"\s*\(.*?\)\s*$", "", str(header)).strip()


def branch_color(branch, generation):
    import colorsys
    if branch == "MEME":
        hue = 345 / 360
    elif branch == "PEPE":
        hue = 205 / 360
    else:
        return None
    try:
        g = int(float(str(generation).strip()))
    except Exception:
        g = 1
    g = max(0, min(g, 5))
    lightness = max(0.26, 0.60 - g * 0.06)
    r, gg, b = colorsys.hls_to_rgb(hue, lightness, 0.55)
    return "#{:02x}{:02x}{:02x}".format(int(r * 255), int(gg * 255), int(b * 255))


def stamp(statut: str) -> str:
    s = (statut or "").strip().lower()
    if is_blank(statut):
        return '<span class="stamp inconnu">à enquêter</span>'
    if s.startswith("confirmé") or s.startswith("confirme"):
        return '<span class="stamp confirme">confirmé</span>'
    return '<span class="stamp hypothese">à confirmer</span>'


def build_fiches(personnes: pd.DataFrame, branches: dict, documents: pd.DataFrame):
    if personnes.empty:
        body = (
            "<p class='empty'>Aucun acteur pour l'instant — dès que le « Registre des personnes » du "
            "tableau Excel contient une première ligne, cette page affichera sa fiche.</p>"
        )
    else:
        idx = by_id_index(personnes)

        def name_or_id(pid):
            pid = str(pid).strip()
            if not pid:
                return None
            row = idx.get(pid)
            return person_display_name(row) if row is not None else pid

        def link_or_text(pid):
            pid = str(pid).strip()
            label = name_or_id(pid)
            if label is None:
                return None
            if pid in idx:
                return f'<a href="#{slug(pid)}">{label}</a>'
            return label

        cards = [
            render_idcard(row, personnes, branches, idx, link_or_text, depth=0, documents=documents)
            for _, row in personnes.iterrows()
        ]
        body = "<h2>Les acteurs de l'histoire, jusqu'ici</h2>" + "".join(cards)
    (DOCS / "fiches.html").write_text(page("Fiches des acteurs", body, active="fiches", depth=0), encoding="utf-8")


def parse_gen(v):
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None


def order_row_with_clusters(row_pids, idx, raw_x):
    """Ordonne une génération en gardant chaque fratrie compacte (jamais coupée par
    quelqu'un d'extérieur), et en collant un conjoint venu d'ailleurs juste à côté
    de son mari/sa femme, à l'extérieur du bloc de fratrie."""
    in_row = set(row_pids)
    parent = {p: p for p in row_pids}

    def find(a):
        while parent[a] != a:
            parent[a] = parent[parent[a]]
            a = parent[a]
        return a

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    for p in row_pids:
        row = idx[p]
        for sib in split_list(row.get(find_col(row, FRATRIE_COLONNES), "")):
            if sib in in_row:
                union(p, sib)

    clusters = {}
    for p in row_pids:
        clusters.setdefault(find(p), []).append(p)

    blocks = []
    singles_with_spouse = []
    for root, members in clusters.items():
        if len(members) == 1:
            p = members[0]
            row = idx[p]
            spouses = [s for s in split_list(row.get(find_col(row, CONJOINT_COLONNES), "")) if s in in_row]
            if spouses and find(spouses[0]) != root:
                singles_with_spouse.append((p, spouses[0]))
                continue
        members_sorted = sorted(members, key=lambda p: raw_x.get(p, 0))
        avg_x = sum(raw_x.get(p, 0) for p in members) / len(members)
        blocks.append([avg_x, members_sorted])
    blocks.sort(key=lambda b: b[0])

    ordered = []
    for _, members in blocks:
        ordered.extend(members)

    for p, spouse in singles_with_spouse:
        if spouse in ordered:
            i = ordered.index(spouse)
            if raw_x.get(p, 0) < raw_x.get(spouse, 0):
                ordered.insert(i, p)
            else:
                ordered.insert(i + 1, p)
        else:
            ordered.append(p)
    return ordered


R = 34          # rayon des cercles
SPACING_X = 132  # écart horizontal entre deux personnes d'une même génération
ROW_H = 178      # écart vertical entre deux générations
MARGIN = 100


def build_index(personnes: pd.DataFrame, branches: dict, meme_id: str, pepe_id: str):
    idx = by_id_index(personnes)

    if not meme_id or personnes.empty:
        body = (
            "<p class='empty'>L'arbre est encore vide — dès que MEME et PEPE sont dans le "
            "« Registre des personnes » (avec « MEME » exactement dans la colonne du lien), "
            "cette page affichera le schéma.</p>"
        )
        (DOCS / "index.html").write_text(page("Accueil", body, active="accueil", depth=0), encoding="utf-8")
        return

    gens = {}
    for pid, row in idx.items():
        g = parse_gen(row.get("Génération", ""))
        if g is not None:
            gens.setdefault(g, []).append(pid)

    base_gen = parse_gen(idx[meme_id].get("Génération", "")) or 1
    positions = {}  # pid -> (x, y)

    def y_for(g):
        return (base_gen - g) * ROW_H

    # -- génération de base : MEME et PEPE --
    base_row = [p for p in (meme_id, pepe_id) if p]
    for i, pid in enumerate(base_row):
        positions[pid] = ((i - (len(base_row) - 1) / 2) * SPACING_X, y_for(base_gen))

    # -- génération des enfants (base_gen - 1), centrée sous le couple --
    child_gen = base_gen - 1
    if child_gen in gens:
        kids = sorted(set(gens[child_gen]))
        center_x = sum(positions[p][0] for p in base_row) / len(base_row) if base_row else 0
        start = center_x - (len(kids) - 1) * SPACING_X / 2
        for i, pid in enumerate(kids):
            positions[pid] = (start + i * SPACING_X, y_for(child_gen))

    # -- générations ancêtres, du plus proche au plus lointain --
    for g in sorted([g for g in gens if g > base_gen]):
        row = sorted(set(gens[g]))
        row_set = set(row)
        raw_x = {}
        for pid in row:
            kids = find_children(pid, personnes)
            kid_xs = [positions[k.get("ID Personne", "").strip()][0]
                      for k in kids if k.get("ID Personne", "").strip() in positions]
            raw_x[pid] = sum(kid_xs) / len(kid_xs) if kid_xs else None
        known_vals = [v for v in raw_x.values() if v is not None]
        left_x = min(known_vals, default=0) - SPACING_X
        right_x = max(known_vals, default=0) + SPACING_X
        for pid in row:
            if raw_x[pid] is None:
                branche = branches.get(pid, "")
                if branche == "MEME":
                    raw_x[pid] = left_x
                    left_x -= SPACING_X
                else:
                    raw_x[pid] = right_x
                    right_x += SPACING_X

        ordered = order_row_with_clusters(row, idx, raw_x)
        center = sum(raw_x.values()) / len(raw_x) if raw_x else 0
        start = center - (len(ordered) - 1) * SPACING_X / 2
        for i, pid in enumerate(ordered):
            positions[pid] = (start + i * SPACING_X, y_for(g))

    if not positions:
        body = "<p class='empty'>Personne n'a encore de génération renseignée.</p>"
        (DOCS / "index.html").write_text(page("Accueil", body, active="accueil", depth=0), encoding="utf-8")
        return

    xs = [p[0] for p in positions.values()]
    ys = [p[1] for p in positions.values()]
    min_x, max_x = min(xs) - MARGIN, max(xs) + MARGIN
    min_y, max_y = min(ys) - MARGIN, max(ys) + MARGIN
    width, height = max_x - min_x, max_y - min_y

    def px(x):
        return x - min_x

    def py(y):
        return y - min_y

    svg_parts = []
    # lignes parent -> enfant
    for pid, (x, y) in positions.items():
        row = idx[pid]
        pere_col = find_col(row, PERE_COLONNES)
        mere_col = find_col(row, MERE_COLONNES)
        for parent_id in (str(row.get(pere_col, "")).strip(), str(row.get(mere_col, "")).strip()):
            if parent_id in positions:
                px1, py1 = positions[parent_id]
                svg_parts.append(
                    f'<line x1="{px(x)}" y1="{py(y)}" x2="{px(px1)}" y2="{py(py1)}" '
                    f'stroke="#c9bda3" stroke-width="2"/>'
                )

    # traits pointillés fins entre frères/sœurs d'une même génération
    by_gen = {}
    for pid, (x, y) in positions.items():
        by_gen.setdefault(y, []).append((x, pid))
    for y, nodes in by_gen.items():
        nodes.sort()
        for (x1, p1), (x2, p2) in zip(nodes, nodes[1:]):
            row1 = idx[p1]
            fratrie1 = split_list(row1.get(find_col(row1, FRATRIE_COLONNES), ""))
            if p2 in fratrie1:
                svg_parts.append(
                    f'<line x1="{px(x1)+R}" y1="{py(y)}" x2="{px(x2)-R}" y2="{py(y)}" '
                    f'stroke="#c9bda3" stroke-width="1.5" stroke-dasharray="2,3"/>'
                )

    # traits en tirets, plus marqués, entre conjoints
    seen_couples = set()
    for pid, (x, y) in positions.items():
        row = idx[pid]
        conjoint_col = find_col(row, CONJOINT_COLONNES)
        for spouse_id in split_list(row.get(conjoint_col, "")):
            if spouse_id in positions and positions[spouse_id][1] == y:
                couple_key = tuple(sorted((pid, spouse_id)))
                if couple_key in seen_couples:
                    continue
                seen_couples.add(couple_key)
                x2, y2 = positions[spouse_id]
                x1s, x2s = sorted((px(x), px(x2)))
                svg_parts.append(
                    f'<line x1="{x1s+R}" y1="{py(y)}" x2="{x2s-R}" y2="{py(y)}" '
                    f'stroke="#8c2f39" stroke-width="2" stroke-dasharray="7,4"/>'
                )

    # noeuds (portrait ou pastille + nom), cliquables
    for pid, (x, y) in positions.items():
        row = idx[pid]
        name = person_display_name(row)
        prenom = row.get("Prénom(s)", "")
        nom = row.get("Nom", "")
        branche = branches.get(pid, "")
        color = branch_color(branche, parse_gen(row.get("Génération", ""))) or "#6b5c48"
        photos = split_urls(row.get("Photos (liens Postimage, séparés par ;)", "")) if "Photos (liens Postimage, séparés par ;)" in row.index else []
        portrait = photos[0] if photos else None
        clip_id = f"clip-{slug(pid)}"
        cx, cy = px(x), py(y)
        if portrait:
            photo_svg = (
                f'<clipPath id="{clip_id}"><circle cx="{cx}" cy="{cy}" r="{R-3}"/></clipPath>'
                f'<image href="{portrait}" x="{cx-R+3}" y="{cy-R+3}" width="{(R-3)*2}" height="{(R-3)*2}" '
                f'preserveAspectRatio="xMidYMid slice" clip-path="url(#{clip_id})"/>'
            )
        else:
            initiales = "".join([w[0] for w in name.split() if w])[:2].upper()
            photo_svg = (
                f'<text x="{cx}" y="{cy+6}" text-anchor="middle" font-family="Playfair Display, Georgia, serif" '
                f'font-size="18" fill="{color}">{initiales}</text>'
            )
        ligne1 = prenom if not is_blank(prenom) else name
        ligne2 = nom if not is_blank(nom) else ""
        svg_parts.append(f"""
        <a href="fiches.html#{slug(pid)}">
          <circle cx="{cx}" cy="{cy}" r="{R}" fill="#faf6ec" stroke="{color}" stroke-width="4"/>
          {photo_svg}
          <text x="{cx}" y="{cy+R+18}" text-anchor="middle" font-family="Public Sans, Arial, sans-serif"
                font-size="12.5" font-weight="700" fill="#3a2e22">{ligne1}</text>
          <text x="{cx}" y="{cy+R+34}" text-anchor="middle" font-family="Public Sans, Arial, sans-serif"
                font-size="12.5" font-weight="700" fill="#3a2e22">{ligne2}</text>
        </a>""")

    svg = (
        f'<svg viewBox="0 0 {width} {height}" width="{width}" height="{height}" '
        f'xmlns="http://www.w3.org/2000/svg">{"".join(svg_parts)}</svg>'
    )
    non_places = len(personnes) - len(positions)
    note = (
        f"<p class='meta'>{non_places} personne(s) sans génération renseignée, donc pas encore sur ce schéma.</p>"
        if non_places > 0 else ""
    )
    body = f"""
    <p class="meta">Cliquez un portrait pour ouvrir sa fiche.</p>
    <div style="overflow-x:auto; background:#faf6ec; border:1px solid var(--line); border-radius:4px; padding:1rem;">
      {svg}
    </div>
    {note}
    """
    (DOCS / "index.html").write_text(page("Accueil", body, active="accueil", depth=0), encoding="utf-8")


def build_person_pages(personnes: pd.DataFrame, branches: dict, documents: pd.DataFrame):
    PERSONNES_DIR.mkdir(parents=True, exist_ok=True)
    idx = by_id_index(personnes)

    def name_or_id(pid):
        pid = str(pid).strip()
        if not pid:
            return None
        row = idx.get(pid)
        return person_display_name(row) if row is not None else pid

    def link_or_text(pid):
        pid = str(pid).strip()
        label = name_or_id(pid)
        if label is None:
            return None
        if pid in idx:
            return f'<a href="{slug(pid)}.html">{label}</a>'
        return label

    for _, row in personnes.iterrows():
        pid = row.get("ID Personne", "")
        name = person_display_name(row)
        card = render_idcard(row, personnes, branches, idx, link_or_text, depth=1, documents=documents)
        body = f"""<a class="back" href="../fiches.html">&larr; Retour aux fiches</a>{card}"""
        (PERSONNES_DIR / f"{slug(pid)}.html").write_text(
            page(name, body, depth=1), encoding="utf-8"
        )


def build_pistes(personnes: pd.DataFrame, documents: pd.DataFrame):
    leads = compute_leads(personnes)
    idx = by_id_index(personnes)
    legend = """
    <div class="legend">
      <span><span class="dot meme"></span>Côté MEME</span>
      <span><span class="dot pepe"></span>Côté PEPE (plus la teinte est foncée, plus la génération est ancienne)</span>
    </div>"""
    if not leads:
        body = legend + "<p class='empty'>Aucune piste ouverte pour l'instant : soit tout est confirmé, soit l'arbre est encore vide.</p>"
    else:
        cards = []
        for lead in leads:
            link = f"fiches.html#{slug(lead['id'])}"
            subject = f"Une piste pour {lead['name']}"
            body_txt = f"Bonjour,%0D%0A%0D%0AÀ propos de {lead['name']} : {lead['text']}%0D%0A%0D%0AVoici ce que j'ai :"
            mailto = f"mailto:{CONTACT_EMAIL}?subject={subject.replace(' ', '%20')}&body={body_txt}"
            docs_html = render_documents_block(lead["id"], documents, idx, title="Ce que l'on a déjà")
            cards.append(f"""
            <div class="card">
              <a class="name" href="{link}">{lead['name']}</a>
              <div class="piste"><strong>Ce qu'il nous manque :</strong> {lead['text']}</div>
              {docs_html}
              <a class="btn" style="display:inline-block; margin-top:.7rem; background:var(--brass); color:#faf6ec;
                 text-decoration:none; padding:.5rem 1rem; border-radius:3px; font-weight:700; font-size:.85rem;"
                 href="{mailto}">Envoyer un document pour cette piste</a>
            </div>""")
        body = legend + "<h2>Ce qu'il nous manque encore</h2>" + "".join(cards)
    (DOCS / "pistes.html").write_text(page("Pistes de recherche", body, active="pistes", depth=0), encoding="utf-8")


def build_archives_privees(personnes: pd.DataFrame, documents: pd.DataFrame):
    idx = by_id_index(personnes)

    def person_label(pid):
        pid = pid.strip()
        row = idx.get(pid)
        label = person_display_name(row) if row is not None else pid
        return f'<a href="fiches.html#{slug(pid)}">{label}</a>' if row is not None else label

    body = f"""
    <h2>Pourquoi certains documents ne sont pas sur ce site</h2>
    <div class="archive-box">
      <p>Les livrets de famille et actes de naissance contiennent des informations
      personnelles qui peuvent concerner des membres de la famille encore vivants
      aujourd'hui (parents, enfants, mentions de mariage ou de reconnaissance).
      Les publier sur un site public ferait courir un vrai risque
      (usurpation d'identité notamment) — on a donc fait le choix de les garder
      strictement privés.</p>
      <p>Ces documents sont conservés sur un espace privé et chiffré
      (Proton Drive), accessible uniquement sur demande.</p>
    </div>
    <h2>Comment en obtenir une copie</h2>
    <p>Écrivez-nous à l'adresse ci-dessous en précisant le document qui vous
    intéresse (ou la personne concernée) : on vous envoie un lien ou un code
    d'accès personnel pour le télécharger.</p>
    """
    if not documents.empty:
        rows = []
        for _, d in documents.iterrows():
            doc_id = d.get("ID Document", "")
            nom = d.get("Nom du doc") or "—"
            lien = get_doc_link(d)
            lien_html = f'<a href="{lien}" target="_blank" rel="noopener">voir</a>' if not is_blank(lien) else "<span class='empty'>pas encore de lien</span>"
            mentionnes = sorted(set(get_personnes_mentionnees(d)))
            noms = [person_label(p) for p in mentionnes]
            lieu = d.get("Rangement physique", "")
            statut = "<span class='field-stamp confirme'>exploité</span>" if mentionnes else "<span class='field-stamp manque'>à explorer</span>"
            rows.append(f"""
            <div class="doc-chip">
              <strong>{doc_id} — {nom}</strong> {statut} · {lien_html}
              {f"<span class='meta'>Chez : {lieu}</span>" if not is_blank(lieu) else ""}
              {f"<span class='meta'>Personnes citées : {', '.join(noms)}</span>" if noms else ""}
            </div>""")
        body += f"<h2>Inventaire de tous les documents scannés ({len(documents)})</h2><div class='doc-list'>{''.join(rows)}</div>"
    (DOCS / "archives-privees.html").write_text(
        page("Archives", body, active="archives", depth=0), encoding="utf-8"
    )


def main():
    DOCS.mkdir(exist_ok=True)
    personnes = load_data()
    documents = load_documents()
    branches, meme_id, pepe_id = compute_branches(personnes)
    build_index(personnes, branches, meme_id, pepe_id)
    build_fiches(personnes, branches, documents)
    build_person_pages(personnes, branches, documents)
    build_pistes(personnes, documents)
    build_archives_privees(personnes, documents)
    print(f"Site généré dans {DOCS} — {len(personnes)} personne(s), {len(documents)} document(s).")


if __name__ == "__main__":
    main()